import pandas as pd
from openpyxl import load_workbook
import os
import re
import unicodedata
import zipfile

# Crear el directorio si no existe
if not os.path.exists('archivos_individuales'):
    os.makedirs('archivos_individuales')

# Cargar el archivo de datos desde la segunda hoja
df_datos = pd.read_excel('/content/Evaluación_de_Desempeño_(Nivel_Estratégico)_(1-22).xlsx', sheet_name=1)

# Normalizar los nombres de las columnas
def normalizar_nombre_columna(nombre):
    nombre = nombre.strip()  # Eliminar espacios en blanco
    nombre = re.sub(r'\s+', '_', nombre)  # Reemplazar espacios por guiones bajos
    nombre = unicodedata.normalize('NFD', nombre)  # Normalizar caracteres acentuados
    nombre = nombre.encode('ascii', 'ignore').decode('utf-8')  # Eliminar acentos
    return nombre

# Aplicar la normalización a todos los nombres de columnas
df_datos.columns = [normalizar_nombre_columna(col) for col in df_datos.columns]

# Omitir la primera columna
df_datos = df_datos.iloc[:, 1:]  # Esto eliminará la primera columna

# Función para limpiar y normalizar valores de celdas
def limpiar_valor(valor):
    if isinstance(valor, str):
        valor = valor.strip()  # Eliminar espacios al principio y al final
        valor = re.sub(r'\s+', ' ', valor)  # Normalizar múltiples espacios a uno solo
        valor = re.sub(r'[^A-Za-z0-9 ÁÉÍÓÚÑáéíóúñ ]', '', valor)  # Permitir solo letras, números y espacios
    return valor

# Función para normalizar niveles
def normalizar_nivel(nivel):
    nivel = nivel.strip()  # Eliminar espacios
    nivel = unicodedata.normalize('NFD', nivel)  # Normalizar caracteres acentuados
    nivel = nivel.encode('ascii', 'ignore').decode('utf-8')  # Eliminar acentos
    return nivel

# Función para marcar el resultado de evaluación
def marcar_resultado_evaluacion(ws, resultadoTotal):
    try:
        resultadoTotal = round(float(resultadoTotal), 2)  # Convertir a float para comparar
        if 4 <= resultadoTotal <= 5:
            ws['D16'] = resultadoTotal
        elif 3 <= resultadoTotal < 4:
            ws['C16'] = resultadoTotal
        else:
            ws['B16'] = resultadoTotal
    except ValueError:
        print(f"Error al convertir el resultado de evaluación: {resultadoTotal}")

# Iterar sobre cada fila (persona) en el DataFrame
for index, row in df_datos.iterrows():
    try:
        # Imprimir fila para debugging
        print(f"Fila {index}: {row.to_dict()}")

        # Trabajador
        nombre_trabajador = limpiar_valor(row['NOMBRE_TRABAJADOR'])
        ID_trabajador = limpiar_valor(row['CEDULA_TRABAJADOR'])  # Asegúrate de que esta columna esté normalizada
        Area_trabajador = limpiar_valor(row['AREA_TRABAJADOR'])  # Ajustar aquí
        cargo_trabajador = limpiar_valor(row['CARGO_TRABAJADOR'])
        nivel_trabajador = normalizar_nivel(row['NIVEL'])  # Normalizar el nivel del trabajador

        # Evaluador
        nombre_evaluador = limpiar_valor(row['NOMBRE_EVALUADOR'])
        ID_evaluador = limpiar_valor(row['CEDULA_EVALUADOR'])  # Asegúrate de que esta columna esté normalizada
        cargo_evaluador = limpiar_valor(row['CARGO_EVALUADOR'])

        # Competencias transversales
        orientacion_resultados = limpiar_valor(row.get('COMPETENCIAS_TRANSVERSALES', 'Sin competencia'))
        gestion_calidad_riesgo = limpiar_valor(row.get('Unnamed:_10', 'Sin competencia'))  # Cambia el nombre de columna
        actitud_servicio = limpiar_valor(row.get('Unnamed:_11', 'Sin competencia'))  # Cambia el nombre de columna
        trabajo_equipo = limpiar_valor(row.get('Unnamed:_12', 'Sin competencia'))  # Cambia el nombre de columna
        comunicacion_asertiva = limpiar_valor(row.get('Unnamed:_13', 'Sin competencia'))  # Cambia el nombre de columna
        adaptabilidad_flexibilidad = limpiar_valor(row.get('Unnamed:_14', 'Sin competencia'))  # Cambia el nombre de columna
        atencion_detalle = limpiar_valor(row.get('Unnamed:_15', 'Sin competencia'))  # Cambia el nombre de columna
        cumplimiento_sst = limpiar_valor(row.get('Unnamed:_16', 'Sin competencia'))  # Cambia el nombre de columna

        # Competencias específicas
        especifica1 = limpiar_valor(row.get('ESPECIFICAS', 'Sin competencia'))  # Cambia el nombre de columna
        especifica2 = limpiar_valor(row.get('Unnamed:_19', 'Sin competencia'))  # Cambia el nombre de columna
        especifica3 = limpiar_valor(row.get('Unnamed:_20', 'Sin competencia'))  # Cambia el nombre de columna

        # Resultado total
        resultadoTotal = limpiar_valor(row.get('RESULTADO_TOTAL_EVALUACION_DE_DESEMPENO', 'Sin competencia'))

        # Verificar si el nivel es NaN
        if pd.isna(nivel_trabajador) or not nivel_trabajador:
            print(f"Fila {index}: NIVEL es NaN o vacío, se omite esta fila.")
            continue

        # Cargar la plantilla
        wb = load_workbook('/content/1242.xlsx')
        hojas = wb.sheetnames

        # Determinar la hoja a editar según el nivel
        if nivel_trabajador.upper() == 'TACTICO':
            ws = wb['RESULTADOS TÁCTICOS']  # Editar la hoja 2
            hojas_a_eliminar = [hoja for hoja in hojas if hoja != 'RESULTADOS TÁCTICOS']  # Hojas a eliminar
        elif nivel_trabajador.upper() == 'OPERATIVO':
            ws = wb['RESULTADOS OPERATIVOS']  # Editar la hoja 1
            hojas_a_eliminar = [hoja for hoja in hojas if hoja != 'RESULTADOS OPERATIVOS']  # Hojas a eliminar
        elif nivel_trabajador.upper() == 'ESTRATEGICO':
            ws = wb['RESULTADOS ESTRATEGICO']  # Editar la hoja 3
            hojas_a_eliminar = [hoja for hoja in hojas if hoja != 'RESULTADOS ESTRATEGICO']  # Hojas a eliminar
        else:
            print(f"Nivel no reconocido: {nivel_trabajador}")
            continue  # Saltar si no se reconoce el nivel

        # Eliminar las hojas que no son necesarias
        for hoja in hojas_a_eliminar:
            wb.remove(wb[hoja]) 

        # Escribir datos en la plantilla
        # Trabajador
        ws['B8'] = nombre_trabajador  
        ws['B9'] = ID_trabajador
        ws['B10'] = Area_trabajador
        ws['D9'] = cargo_trabajador
        ws['D10'] = nivel_trabajador

        # Evaluador
        ws['B12'] = nombre_evaluador
        ws['B13'] = ID_evaluador
        ws['D13'] = cargo_evaluador

        # Competencias transversales
        ws['D18'] = orientacion_resultados
        ws['D19'] = gestion_calidad_riesgo
        ws['D20'] = actitud_servicio
        ws['D21'] = trabajo_equipo
        ws['D22'] = comunicacion_asertiva
        ws['D23'] = adaptabilidad_flexibilidad
        ws['D24'] = atencion_detalle
        ws['D25'] = cumplimiento_sst

        # Competencias específicas
        ws['D27'] = especifica1
        ws['D28'] = especifica2
        ws['D29'] = especifica3

        # Marcar el resultado total de la evaluación
        marcar_resultado_evaluacion(ws, resultadoTotal)

        # Guardar el archivo con un nombre único
        nombre_archivo = f'archivos_individuales/{nombre_trabajador}.xlsx'
        wb.save(nombre_archivo)
        wb.close()

    except Exception as e:
        print(f"Error en la fila {index}: {e}")

with zipfile.ZipFile('archivos_evaluaciones.zip', 'w') as zipf:
    for root, dirs, files in os.walk('archivos_individuales'):
        for file in files:
            zipf.write(os.path.join(root, file))

print("Archivos individuales generados exitosamente con formato.")